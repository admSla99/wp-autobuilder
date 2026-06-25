#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
compile_rules.py — prevedie content-mapping Excel (01_SABLONA_v2.xlsx)
na normalizovany rules.json a zvaliduje ho.

Pouzitie:
    python compile_rules.py [cesta_k_xlsx] [-o rules.json]

Vystup: rules.json + validacny report na stdout.
Navratovy kod: 0 = OK (aj s varovaniami), 1 = chyby (nepustaj tvorbu stranky).
"""
import sys, os, re, json, unicodedata, argparse, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Chyba: nainstaluj openpyxl (pip install openpyxl).")

# --- sheet -> (page_key, elementor_target) ---
PAGES = {
    'HP': ('hp', 'Hp'),
    'o nás': ('about', 'O nás'),
    'SLuzby': ('services', 'Služby'),
    'kontakt': ('contact', 'Kontakt'),
    'cenník': ('pricing', 'cenník'),
    'produkty': ('products', 'Produktova stránka'),
    'PRODUKTOVA STRANKA': ('product_page', 'Produktova stránka'),
    'galeria': ('gallery', 'Jednoúrovňová galéria'),
    'blogový článok': ('blog', 'Blog – článok'),
    'vizitka,vlastné referenice': ('team_refs', 'Referencie / Tím'),
    'okolie,výlety,zážitky': ('area', 'Okolie / Výlety / Zážitky'),
}

GLOBAL_INVARIANTS = [
    "Nevymýšľať ceny — vždy ich dodá klient.",
    "Nevymýšľať recenzie, mená ani hodnotenia.",
    "Nemeniť fakty, čísla, legislatívu, technické parametre ani poradie fotiek.",
    "Nemeniť brand farby, fonty, layout ani štruktúru menu.",
    "CTA default max 21 znakov; nepoužívať „Kliknite sem / Viac / Tu / Kúpte hneď\".",
]

TOOL_ROUTING = {
    "foto": "Image 2.0 (cez n8n)",
    "video": "Runway (cez n8n)",
    "text": "Claude",
    "recenzie": "Google Reviews API",
    "podklady_klienta": "Google Drive",
}

SEC_ID = re.compile(r'^(S|V\d+|\d+(\.\d+)?|\d+[A-Z])$')
MARKERS = ('▶', '★', 'ℹ', '⚠', '✓', 'LEGENDA', 'PRVOK')


def norm(v):
    return (unicodedata.normalize('NFKD', str(v or ''))
            .encode('ascii', 'ignore').decode().lower().strip())


def map_header(label):
    n = norm(label)
    if n == '#':
        return 'order'
    if n == 'stav':
        return 'stav'
    if n.startswith('nazov sekcie'):
        return 'name'
    if n.startswith('typ prvkov'):
        return 'element_types'
    if n.startswith('pocet poloziek'):
        return 'item_count'
    if n.startswith('vizual'):
        return 'visual_type'
    if n.startswith('pocet vizual'):
        return 'visual_count'
    if n.startswith('ai moze menit'):
        return 'ai_can_edit'
    if n.startswith('ai nesmie menit') or n.startswith('ai nemoze menit') or n.startswith('ai nemezo'):
        return 'ai_cannot_edit'
    if n.startswith('zdroj dat'):
        return 'data_source'
    if n.startswith('podmienky'):
        return 'rules_text'
    if n == 'slot_id':
        return 'slot_id'
    return None


def parse_count(text):
    if text is None:
        return {"raw": None, "min": None, "max": None, "unlimited": False}
    raw = str(text).strip()
    low = norm(raw)
    unlimited = ('neobmedzen' in low) or ('bez obmedzen' in low) or ('bez limitu' in low)
    m = re.search(r'min\.?\s*(\d+)\D+max\.?\s*(\d+)', low)
    if m:
        return {"raw": raw, "min": int(m.group(1)), "max": int(m.group(2)), "unlimited": unlimited}
    m = re.search(r'(\d+)\s*[-–]\s*(\d+)', raw)
    if m:
        return {"raw": raw, "min": int(m.group(1)), "max": int(m.group(2)), "unlimited": unlimited}
    m = re.search(r'(\d+)\s*\+', raw)
    if m:
        return {"raw": raw, "min": int(m.group(1)), "max": None, "unlimited": True}
    nums = re.findall(r'(\d+)', raw)
    if nums and not unlimited:
        n = int(nums[0])
        return {"raw": raw, "min": n, "max": n, "unlimited": False}
    return {"raw": raw, "min": None, "max": None, "unlimited": unlimited}


def split_list(text):
    if not text:
        return []
    parts = re.split(r'[;,]\s*|\s*\+\s*|\s*/\s*', str(text))
    return [p.strip() for p in parts if p.strip() and p.strip() != '—']


def parse_constraints(name, rules_text, item_count):
    c = {}
    blob = " ".join(filter(None, [str(name or ''), str(rules_text or '')]))
    low = norm(blob)
    m = re.search(r'cta[^.]*?max\.?\s*(\d+)\s*znak', low)
    if not m:
        m = re.search(r'max\.?\s*(\d+)\s*znak', low)
    if m:
        c['cta_max_chars'] = int(m.group(1))
    if 'cen' in low and ('vymysl' in low or 'klient musi dodat' in low or 'klient dodava: nazvy, ceny' in low):
        c['price_client_required'] = True
    if 'nevym' in low and ('recenz' in low or 'mena' in low or 'hodnoten' in low):
        c['no_invented_reviews'] = True
    if 'porad' in low and 'foto' in low and ('nesmie' in low or 'nemeni' in low):
        c['photo_order_locked'] = True
    if 'foto + video' in low or ('1 foto' in low and '1 video' in low) or 'foto + 1 video' in low:
        c['pair_photo_video'] = True
    if 'akordeon' in low and ('legislat' in low or 'postup' in low or 'nesmie byt obsahovo zmen' in low):
        c['preserve_verbatim'] = True
    return c


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', nargs='?', default=None)
    ap.add_argument('-o', '--out', default=None)
    args = ap.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    xlsx = args.xlsx or os.path.join(here, '01_SABLONA_v2.xlsx')
    out = args.out or os.path.join(here, 'rules.json')
    if not os.path.exists(xlsx):
        sys.exit(f"Chyba: nenajdeny subor {xlsx}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    errors, warnings = [], []
    pages = {}

    for ws in wb.worksheets:
        if ws.title not in PAGES:
            continue
        pk, target = PAGES[ws.title]
        sheet_max_col = 16
        # header rows
        header_rows = [r for r in range(1, ws.max_row + 1)
                       if str(ws.cell(r, 1).value or '').strip() == '#']
        # intro notes = non-empty rows above first header
        intro = []
        first_h = header_rows[0] if header_rows else ws.max_row + 1
        for r in range(1, first_h):
            vals = [str(ws.cell(r, c).value).strip() for c in range(1, sheet_max_col)
                    if ws.cell(r, c).value not in (None, '')]
            if vals:
                intro.append(" | ".join(vals))

        sections = []
        seen_slots = {}
        for hr in header_rows:
            colmap = {}
            for c in range(1, sheet_max_col):
                key = map_header(ws.cell(hr, c).value)
                if key and key not in colmap:
                    colmap[key] = c
            if 'name' not in colmap:
                warnings.append(f"[{pk}] hlavicka v riadku {hr} nema stlpec 'Nazov sekcie'")
                continue
            r = hr + 1
            while r <= ws.max_row:
                a = ws.cell(r, 1).value
                nm = ws.cell(r, colmap['name']).value
                if a is None and ws.cell(r, 2).value is None:
                    break
                if str(a or '').strip() == '#':
                    break
                aval = str(a or '').strip()
                if aval and not aval.startswith(MARKERS) and SEC_ID.match(aval) and nm not in (None, ''):
                    def g(k):
                        return ws.cell(r, colmap[k]).value if k in colmap else None
                    slot = (str(g('slot_id')).strip() if g('slot_id') else None)
                    item_count = parse_count(g('item_count'))
                    rules_text = g('rules_text')
                    sec = {
                        "order": aval,
                        "slot_id": slot,
                        "stav": (str(g('stav')).strip().upper() if g('stav') else "POVINNÁ"),
                        "name": str(nm).replace('\n', ' ').strip(),
                        "element_types": (str(g('element_types')).replace('\n', ' ').strip()
                                          if g('element_types') else None),
                        "item_count": item_count,
                        "visual_type": (str(g('visual_type')).replace('\n', ' ').strip()
                                        if g('visual_type') else None),
                        "visual_count": (str(g('visual_count')).strip() if g('visual_count') else None),
                        "ai_can_edit": split_list(g('ai_can_edit')),
                        "ai_cannot_edit": split_list(g('ai_cannot_edit')),
                        "data_source": split_list(g('data_source')),
                        "rules_text": (str(rules_text).replace('\n', ' ').strip()
                                       if rules_text else None),
                        "constraints": parse_constraints(nm, rules_text, item_count),
                    }
                    # validations
                    if not slot:
                        errors.append(f"[{pk}] sekcia '{sec['name']}' (#{aval}) nema slot_id")
                    elif slot in seen_slots:
                        errors.append(f"[{pk}] duplicitny slot_id '{slot}'")
                    else:
                        seen_slots[slot] = True
                    if item_count['raw'] and item_count['min'] is None and not item_count['unlimited']:
                        warnings.append(f"[{pk}/{slot}] nečitateľný počet: '{item_count['raw']}'")
                    vt = norm(sec['visual_type'])
                    if vt and vt not in ('nie', '—', '', '-') and sec['visual_count'] in (None, '—', '-', ''):
                        warnings.append(f"[{pk}/{slot}] vizuál '{sec['visual_type']}' bez počtu vizuálov")
                    sections.append(sec)
                r += 1

        if not sections:
            warnings.append(f"[{pk}] ziadne sekcie nenajdene")
        pages[pk] = {
            "sheet": ws.title,
            "elementor_target": target,
            "intro_notes": intro,
            "sections": sections,
        }

    doc = {
        "$schema": "./rules.schema.json",
        "version": "2.0",
        "source_file": os.path.basename(xlsx),
        "generated_at": datetime.datetime.now().isoformat(timespec='seconds'),
        "global_invariants": GLOBAL_INVARIANTS,
        "tool_routing": TOOL_ROUTING,
        "pages": pages,
        "stats": {
            "pages": len(pages),
            "sections_total": sum(len(p["sections"]) for p in pages.values()),
            "errors": len(errors),
            "warnings": len(warnings),
        },
    }
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)

    print("=" * 60)
    print(f"COMPILE RULES — {os.path.basename(xlsx)}")
    print(f"Stránok: {doc['stats']['pages']} | Sekcií: {doc['stats']['sections_total']}")
    print(f"Výstup: {out}")
    print("-" * 60)
    if errors:
        print(f"❌ CHYBY ({len(errors)}):")
        for e in errors:
            print("   -", e)
    if warnings:
        print(f"⚠ VAROVANIA ({len(warnings)}):")
        for w in warnings[:40]:
            print("   -", w)
        if len(warnings) > 40:
            print(f"   ... (+{len(warnings)-40})")
    if not errors and not warnings:
        print("✅ Všetko v poriadku.")
    print("=" * 60)
    sys.exit(1 if errors else 0)


if __name__ == '__main__':
    main()
